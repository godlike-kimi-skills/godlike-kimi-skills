#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel Processor - A comprehensive tool for Excel file manipulation
Supports reading, writing, formatting, formulas, and charts

Author: godlike-kimi
License: MIT
"""

import argparse
import json
import os
import sys
from typing import Any, Dict, List, Optional, Union
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelProcessor:
    """Main class for Excel file processing"""

    def __init__(self, file_path: Optional[str] = None):
        self.file_path = file_path
        self.workbook = None
        if file_path and os.path.exists(file_path):
            self.workbook = load_workbook(file_path)
        else:
            self.workbook = Workbook()

    def get_sheet(self, sheet_name: Optional[str] = None) -> Any:
        """Get worksheet by name or return active sheet"""
        if sheet_name:
            if sheet_name in self.workbook.sheetnames:
                return self.workbook[sheet_name]
            else:
                return self.workbook.create_sheet(sheet_name)
        return self.workbook.active

    def read(self, sheet_name: Optional[str] = None, 
             cell_range: Optional[str] = None,
             headers: bool = True) -> List[Dict[str, Any]]:
        """Read data from Excel file"""
        sheet = self.get_sheet(sheet_name)
        data = []

        if cell_range:
            # Parse range like "A1:D10"
            cells = sheet[cell_range]
            if headers:
                header_row = [cell.value for cell in cells[0]]
                for row in cells[1:]:
                    row_data = {}
                    for idx, cell in enumerate(row):
                        if idx < len(header_row):
                            row_data[header_row[idx]] = cell.value
                    data.append(row_data)
            else:
                for row in cells:
                    data.append([cell.value for cell in row])
        else:
            # Read all data
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return data

            if headers:
                header_row = rows[0]
                for row in rows[1:]:
                    row_data = {}
                    for idx, value in enumerate(row):
                        if idx < len(header_row):
                            row_data[header_row[idx]] = value
                    data.append(row_data)
            else:
                data = rows

        return data

    def write(self, data: Union[List[Dict], List[List]], 
              sheet_name: Optional[str] = None,
              headers: Union[bool, List[str]] = True,
              start_cell: str = "A1") -> None:
        """Write data to Excel file"""
        sheet = self.get_sheet(sheet_name)
        start_row = int(''.join(filter(str.isdigit, start_cell)))
        start_col = ''.join(filter(str.isalpha, start_cell))
        start_col_idx = self._col_letter_to_idx(start_col)

        if not data:
            return

        # Determine headers
        header_list = []
        if isinstance(headers, list):
            header_list = headers
        elif headers and isinstance(data[0], dict):
            header_list = list(data[0].keys())

        # Write headers
        current_row = start_row
        if header_list:
            for idx, header in enumerate(header_list):
                col_letter = get_column_letter(start_col_idx + idx)
                sheet[f'{col_letter}{current_row}'] = header
            current_row += 1

        # Write data
        for row_data in data:
            if isinstance(row_data, dict):
                for idx, key in enumerate(header_list):
                    col_letter = get_column_letter(start_col_idx + idx)
                    sheet[f'{col_letter}{current_row}'] = row_data.get(key)
            else:
                for idx, value in enumerate(row_data):
                    col_letter = get_column_letter(start_col_idx + idx)
                    sheet[f'{col_letter}{current_row}'] = value
            current_row += 1

        # Auto-adjust column widths
        self._auto_adjust_columns(sheet)

    def append(self, data: Union[Dict, List], 
               sheet_name: Optional[str] = None) -> None:
        """Append a row to existing sheet"""
        sheet = self.get_sheet(sheet_name)
        next_row = sheet.max_row + 1

        if isinstance(data, dict):
            # Get headers from first row
            headers = [cell.value for cell in sheet[1]]
            for idx, header in enumerate(headers):
                col_letter = get_column_letter(idx + 1)
                sheet[f'{col_letter}{next_row}'] = data.get(header)
        else:
            for idx, value in enumerate(data):
                col_letter = get_column_letter(idx + 1)
                sheet[f'{col_letter}{next_row}'] = value

    def merge_files(self, files: List[str], output: str, 
                    merge_type: str = "append") -> None:
        """Merge multiple Excel files"""
        merged_wb = Workbook()
        merged_wb.remove(merged_wb.active)  # Remove default sheet

        if merge_type == "append":
            # Stack files vertically
            target_sheet = merged_wb.create_sheet("Merged")
            current_row = 1
            first_file = True

            for file_path in files:
                if not os.path.exists(file_path):
                    continue

                wb = load_workbook(file_path)
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                        if not first_file and row_idx == 1:
                            continue  # Skip headers for subsequent files
                        for col_idx, value in enumerate(row, 1):
                            target_sheet.cell(row=current_row, column=col_idx, value=value)
                        current_row += 1
                    first_file = False
        else:
            # Side by side
            for file_path in files:
                if not os.path.exists(file_path):
                    continue
                wb = load_workbook(file_path)
                for sheet_name in wb.sheetnames:
                    source_sheet = wb[sheet_name]
                    target_sheet = merged_wb.create_sheet(title=f"{os.path.basename(file_path)}_{sheet_name}")
                    for row_idx, row in enumerate(source_sheet.iter_rows(values_only=True), 1):
                        for col_idx, value in enumerate(row, 1):
                            target_sheet.cell(row=row_idx, column=col_idx, value=value)

        merged_wb.save(output)

    def format_cells(self, cell_range: str, sheet_name: Optional[str] = None,
                     font: Optional[Dict] = None,
                     fill: Optional[Dict] = None,
                     alignment: Optional[Dict] = None,
                     border: Optional[Dict] = None) -> None:
        """Apply formatting to cells"""
        sheet = self.get_sheet(sheet_name)

        # Create styles
        cell_font = None
        if font:
            cell_font = Font(
                name=font.get('name', 'Arial'),
                size=font.get('size', 11),
                bold=font.get('bold', False),
                italic=font.get('italic', False),
                color=font.get('color', '000000')
            )

        cell_fill = None
        if fill:
            cell_fill = PatternFill(
                start_color=fill.get('color', 'FFFFFF'),
                end_color=fill.get('color', 'FFFFFF'),
                fill_type=fill.get('type', 'solid')
            )

        cell_alignment = None
        if alignment:
            cell_alignment = Alignment(
                horizontal=alignment.get('horizontal', 'left'),
                vertical=alignment.get('vertical', 'center'),
                wrap_text=alignment.get('wrap_text', False)
            )

        cell_border = None
        if border:
            side = Side(
                style=border.get('style', 'thin'),
                color=border.get('color', '000000')
            )
            cell_border = Border(left=side, right=side, top=side, bottom=side)

        # Apply to range
        for row in sheet[cell_range]:
            for cell in row:
                if cell_font:
                    cell.font = cell_font
                if cell_fill:
                    cell.fill = cell_fill
                if cell_alignment:
                    cell.alignment = cell_alignment
                if cell_border:
                    cell.border = cell_border

    def add_formula(self, cell: str, formula: str, 
                    sheet_name: Optional[str] = None) -> None:
        """Add formula to a cell"""
        sheet = self.get_sheet(sheet_name)
        sheet[cell] = formula

    def create_chart(self, chart_type: str, data_range: str, 
                     title: str, sheet_name: Optional[str] = None,
                     target_sheet: Optional[str] = None) -> None:
        """Create a chart in the workbook"""
        source_sheet = self.get_sheet(sheet_name)
        dest_sheet = self.get_sheet(target_sheet) if target_sheet else source_sheet

        # Create appropriate chart type
        if chart_type == 'bar':
            chart = BarChart()
        elif chart_type == 'line':
            chart = LineChart()
        elif chart_type == 'pie':
            chart = PieChart()
        elif chart_type == 'scatter':
            chart = ScatterChart()
        else:
            chart = BarChart()

        chart.title = title
        chart.style = 10

        # Parse data range
        data = Reference(source_sheet, min_col=1, min_row=1, max_col=4, max_row=10)
        categories = Reference(source_sheet, min_col=1, min_row=2, max_row=10)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        # Add chart to sheet
        dest_sheet.add_chart(chart, "E2")

    def save(self, output_path: Optional[str] = None) -> None:
        """Save the workbook"""
        path = output_path or self.file_path
        if not path:
            raise ValueError("No output path specified")
        self.workbook.save(path)

    def _col_letter_to_idx(self, col_letter: str) -> int:
        """Convert column letter to index (A=1, B=2, ...)"""
        result = 0
        for char in col_letter.upper():
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result

    def _auto_adjust_columns(self, sheet) -> None:
        """Auto-adjust column widths based on content"""
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width


def main():
    """Command line interface"""
    parser = argparse.ArgumentParser(
        description='Excel Processor - Comprehensive Excel manipulation tool',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python main.py read --input data.xlsx --sheet Sheet1
  python main.py write --input data.xlsx --data '[{"name":"John","age":30}]'
  python main.py format --input data.xlsx --range A1:D10 --fill '{"color":"FFFF00"}'
  python main.py chart --input data.xlsx --chart_type bar --title "Sales Chart"
        """
    )

    parser.add_argument('action', choices=['read', 'write', 'append', 'merge', 'format', 'chart', 'formula'],
                        help='Action to perform')
    parser.add_argument('--input', '-i', help='Input file path')
    parser.add_argument('--output', '-o', help='Output file path')
    parser.add_argument('--sheet', '-s', default='Sheet1', help='Sheet name')
    parser.add_argument('--data', '-d', help='JSON data string')
    parser.add_argument('--range', '-r', help='Cell range (e.g., A1:D10)')
    parser.add_argument('--headers', type=lambda x: x.lower() == 'true', default=True,
                        help='Include headers (true/false)')
    parser.add_argument('--chart_type', choices=['bar', 'line', 'pie', 'scatter'],
                        help='Chart type for chart action')
    parser.add_argument('--title', '-t', help='Chart title or formula')
    parser.add_argument('--font', help='Font settings as JSON')
    parser.add_argument('--fill', help='Fill settings as JSON')
    parser.add_argument('--alignment', help='Alignment settings as JSON')
    parser.add_argument('--formula', '-f', help='Excel formula')
    parser.add_argument('--cell', '-c', help='Cell reference for formula')
    parser.add_argument('--files', help='Comma-separated list of files for merge')

    args = parser.parse_args()

    processor = ExcelProcessor(args.input)

    try:
        if args.action == 'read':
            data = processor.read(args.sheet, args.range, args.headers)
            print(json.dumps(data, indent=2, ensure_ascii=False))

        elif args.action == 'write':
            if not args.data:
                print("Error: --data is required for write action", file=sys.stderr)
                sys.exit(1)
            data = json.loads(args.data)
            if not isinstance(data, list):
                data = [data]
            processor.write(data, args.sheet, args.headers)
            processor.save(args.output or args.input)
            print(f"Data written successfully to {args.output or args.input}")

        elif args.action == 'append':
            if not args.data:
                print("Error: --data is required for append action", file=sys.stderr)
                sys.exit(1)
            data = json.loads(args.data)
            processor.append(data, args.sheet)
            processor.save(args.output or args.input)
            print(f"Data appended successfully")

        elif args.action == 'merge':
            if not args.files:
                print("Error: --files is required for merge action", file=sys.stderr)
                sys.exit(1)
            files = args.files.split(',')
            processor.merge_files(files, args.output or 'merged.xlsx')
            print(f"Files merged successfully to {args.output or 'merged.xlsx'}")

        elif args.action == 'format':
            if not args.range:
                print("Error: --range is required for format action", file=sys.stderr)
                sys.exit(1)
            font = json.loads(args.font) if args.font else None
            fill = json.loads(args.fill) if args.fill else None
            alignment = json.loads(args.alignment) if args.alignment else None
            processor.format_cells(args.range, args.sheet, font, fill, alignment)
            processor.save(args.output or args.input)
            print(f"Formatting applied successfully")

        elif args.action == 'chart':
            if not args.chart_type:
                print("Error: --chart_type is required for chart action", file=sys.stderr)
                sys.exit(1)
            processor.create_chart(args.chart_type, args.range or 'A1:D10',
                                   args.title or 'Chart', args.sheet)
            processor.save(args.output or args.input)
            print(f"Chart created successfully")

        elif args.action == 'formula':
            if not args.cell or not args.formula:
                print("Error: --cell and --formula are required", file=sys.stderr)
                sys.exit(1)
            processor.add_formula(args.cell, args.formula, args.sheet)
            processor.save(args.output or args.input)
            print(f"Formula added successfully")

    except Exception as e:
        print(f"Error: {str(e)}", file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()
